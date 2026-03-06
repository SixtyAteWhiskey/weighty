from __future__ import annotations

import os
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from flask import Flask, flash, g, redirect, render_template, request, url_for
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / "tracker.db"
DEFAULT_XLSX_CANDIDATES = [
    BASE_DIR / "Weight Loss Trajectory.xlsx",
    DATA_DIR / "Weight Loss Trajectory.xlsx",
    Path("/app/Weight Loss Trajectory.xlsx"),
    Path("/app/data/Weight Loss Trajectory.xlsx"),
]

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-change-me")

ACTIVITY_MULTIPLIERS = {
    "Sedentary": 1.2,
    "Lightly Active": 1.3,
    "Moderately Active": 1.5,
    "Intensely Active": 1.7,
}
DEFAULT_ACTIVITY_LEVEL = "Moderately Active"


@dataclass
class Profile:
    age: int
    start_date: date
    start_weight: float
    height_inches: float
    calorie_goal: int
    activity_level: str


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(_: Any) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    db = sqlite3.connect(DB_PATH)
    cur = db.cursor()
    cur.executescript(
        """
        CREATE TABLE IF NOT EXISTS profile (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            age INTEGER NOT NULL,
            start_date TEXT NOT NULL,
            start_weight REAL NOT NULL,
            height_inches REAL NOT NULL,
            calorie_goal INTEGER NOT NULL,
            activity_level TEXT NOT NULL DEFAULT 'Moderately Active'
        );

        CREATE TABLE IF NOT EXISTS weekly_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            week_start TEXT NOT NULL UNIQUE,
            actual_weight REAL,
            weekly_calories INTEGER,
            note TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS daily_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_date TEXT NOT NULL UNIQUE,
            weight REAL,
            calories INTEGER,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );
        """
    )

    existing_columns = {row[1] for row in cur.execute("PRAGMA table_info(profile)").fetchall()}
    if "activity_level" not in existing_columns:
        cur.execute(
            "ALTER TABLE profile ADD COLUMN activity_level TEXT NOT NULL DEFAULT 'Moderately Active'"
        )
        cur.execute(
            "UPDATE profile SET activity_level = ? WHERE activity_level IS NULL OR activity_level = ''",
            (DEFAULT_ACTIVITY_LEVEL,),
        )

    db.commit()
    db.close()


def calculate_bmi(weight_lbs: float | None, height_inches: float) -> float | None:
    if weight_lbs in (None, 0) or height_inches in (None, 0):
        return None
    return round((703 * weight_lbs) / (height_inches ** 2), 1)


def calculate_bmr(weight_lbs: float | None, height_inches: float, age: int) -> float | None:
    if weight_lbs is None:
        return None
    return round(4.38 * weight_lbs + 14.55 * height_inches - 5.08 * age + 260, 0)


def week_start_for(dt: date) -> date:
    return dt - timedelta(days=dt.weekday())


def get_profile() -> Profile | None:
    row = get_db().execute("SELECT * FROM profile WHERE id = 1").fetchone()
    if not row:
        return None
    activity_level = row["activity_level"] if row["activity_level"] in ACTIVITY_MULTIPLIERS else DEFAULT_ACTIVITY_LEVEL
    return Profile(
        age=row["age"],
        start_date=date.fromisoformat(row["start_date"]),
        start_weight=float(row["start_weight"]),
        height_inches=float(row["height_inches"]),
        calorie_goal=int(row["calorie_goal"]),
        activity_level=activity_level,
    )


def seed_profile_from_defaults() -> None:
    db = get_db()
    existing = db.execute("SELECT 1 FROM profile WHERE id = 1").fetchone()
    if existing:
        return

    seeded = False
    for xlsx_path in DEFAULT_XLSX_CANDIDATES:
        if xlsx_path.exists():
            try:
                seeded = import_from_excel(db, xlsx_path)
                if seeded:
                    break
            except Exception as exc:
                print(f"Excel import failed for {xlsx_path}: {exc}")

    if not seeded:
        today = date.today()
        db.execute(
            """
            INSERT OR REPLACE INTO profile (id, age, start_date, start_weight, height_inches, calorie_goal, activity_level)
            VALUES (1, ?, ?, ?, ?, ?, ?)
            """,
            (27, today.isoformat(), 260, 76, 2100, DEFAULT_ACTIVITY_LEVEL),
        )
        db.commit()


def import_from_excel(db: sqlite3.Connection, path: Path) -> bool:
    wb = load_workbook(path, data_only=True)
    overview = wb["Overview"]
    weekly = wb["Weekly CalWeigh in"]

    age = int(overview["C2"].value)
    start_date_value = overview["C3"].value
    start_date = start_date_value.date() if isinstance(start_date_value, datetime) else start_date_value
    start_weight = float(overview["C4"].value)
    height_inches = float(overview["C6"].value)
    calorie_goal = int(overview["C7"].value)

    existing_profile = db.execute("SELECT activity_level FROM profile WHERE id = 1").fetchone()
    activity_level = (
        existing_profile["activity_level"]
        if existing_profile and existing_profile["activity_level"] in ACTIVITY_MULTIPLIERS
        else DEFAULT_ACTIVITY_LEVEL
    )

    db.execute(
        """
        INSERT OR REPLACE INTO profile (id, age, start_date, start_weight, height_inches, calorie_goal, activity_level)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (1, age, start_date.isoformat(), start_weight, height_inches, calorie_goal, activity_level),
    )

    for row in range(12, 65):
        week_date = overview.cell(row=row, column=2).value
        actual_weight = overview.cell(row=row, column=5).value
        weekly_calories = overview.cell(row=row, column=12).value
        if isinstance(week_date, datetime):
            week_date = week_date.date()
        if not week_date:
            continue
        if actual_weight is None and weekly_calories is None:
            continue
        db.execute(
            """
            INSERT OR IGNORE INTO weekly_entries (week_start, actual_weight, weekly_calories, note)
            VALUES (?, ?, ?, ?)
            """,
            (
                week_date.isoformat(),
                float(actual_weight) if actual_weight is not None else None,
                int(weekly_calories) if weekly_calories is not None else None,
                "Imported from workbook",
            ),
        )

    start_current = weekly["B2"].value
    if isinstance(start_current, datetime):
        start_current = start_current.date()

    for col in range(2, 9):
        day_date = weekly.cell(row=2, column=col).value
        if isinstance(day_date, datetime):
            day_date = day_date.date()
        if not day_date and start_current:
            day_date = start_current + timedelta(days=col - 2)
        if not day_date:
            continue
        weight = weekly.cell(row=4, column=col).value
        calories = weekly.cell(row=5, column=col).value
        if weight is None and calories is None:
            continue
        db.execute(
            """
            INSERT OR IGNORE INTO daily_entries (entry_date, weight, calories)
            VALUES (?, ?, ?)
            """,
            (
                day_date.isoformat(),
                float(weight) if weight is not None else None,
                int(calories) if calories is not None else None,
            ),
        )

    db.commit()
    return True


@app.before_request
def ensure_seeded() -> None:
    init_db()
    seed_profile_from_defaults()


@app.context_processor
def inject_helpers() -> dict[str, Any]:
    return {"today": date.today(), "activity_multipliers": ACTIVITY_MULTIPLIERS}


@app.route("/")
def index() -> str:
    profile = get_profile()
    if not profile:
        raise RuntimeError("Profile not initialized")

    db = get_db()
    weekly_rows = db.execute(
        "SELECT id, week_start, actual_weight, weekly_calories, note FROM weekly_entries ORDER BY week_start ASC"
    ).fetchall()
    daily_rows = db.execute(
        "SELECT id, entry_date, weight, calories FROM daily_entries ORDER BY entry_date ASC"
    ).fetchall()

    actual_by_week = {date.fromisoformat(r["week_start"]): r for r in weekly_rows}
    today = date.today()
    current_week_start = week_start_for(today)

    latest_weight_row = next((r for r in reversed(daily_rows) if r["weight"] is not None), None)
    current_weight = float(latest_weight_row["weight"]) if latest_weight_row else profile.start_weight
    current_bmi = calculate_bmi(current_weight, profile.height_inches)
    current_bmr = calculate_bmr(current_weight, profile.height_inches, profile.age)
    selected_activity_multiplier = ACTIVITY_MULTIPLIERS.get(profile.activity_level, ACTIVITY_MULTIPLIERS[DEFAULT_ACTIVITY_LEVEL])
    selected_tdee = round(current_bmr * selected_activity_multiplier, 0) if current_bmr is not None else None

    avg_daily_calories = None
    daily_cal_values = [r["calories"] for r in daily_rows if r["calories"] is not None]
    if daily_cal_values:
        avg_daily_calories = round(sum(daily_cal_values) / len(daily_cal_values), 0)

    estimated_daily_deficit = None
    estimated_weekly_loss = None
    if selected_tdee is not None and avg_daily_calories is not None:
        estimated_daily_deficit = round(selected_tdee - avg_daily_calories, 0)
        estimated_weekly_loss = round((estimated_daily_deficit * 7) / 3500, 2)

    projection = []
    for week_index in range(0, 26):
        week_date = profile.start_date + timedelta(days=7 * week_index)
        expected_weight = round(profile.start_weight - (2 * week_index), 1)
        actual_row = actual_by_week.get(week_date)
        actual_weight = float(actual_row["actual_weight"]) if actual_row and actual_row["actual_weight"] is not None else None
        current_projection_weight = actual_weight if actual_weight is not None else expected_weight
        projection_bmr = calculate_bmr(current_projection_weight, profile.height_inches, profile.age)
        tdees = {
            name: round(projection_bmr * mult, 0) if projection_bmr is not None else None
            for name, mult in ACTIVITY_MULTIPLIERS.items()
        }
        weekly_cals = int(actual_row["weekly_calories"]) if actual_row and actual_row["weekly_calories"] is not None else None
        deficits = {
            name: (round(7 * val - weekly_cals, 0) if weekly_cals is not None and val is not None else None)
            for name, val in tdees.items()
        }
        fat_loss = {
            name: (round(deficit / 3500, 2) if deficit is not None else None)
            for name, deficit in deficits.items()
        }

        projection.append(
            {
                "week_date": week_date,
                "expected_weight": expected_weight,
                "expected_bmi": calculate_bmi(expected_weight, profile.height_inches),
                "actual_weight": actual_weight,
                "actual_bmi": calculate_bmi(actual_weight, profile.height_inches) if actual_weight is not None else None,
                "current_bmr": projection_bmr,
                "tdees": tdees,
                "weekly_cals": weekly_cals,
                "deficits": deficits,
                "fat_loss": fat_loss,
                "selected_fat_loss": fat_loss.get(profile.activity_level),
                "is_current": week_date == current_week_start,
            }
        )

    chart_labels = [r["entry_date"] for r in daily_rows if r["weight"] is not None]
    chart_weights = [float(r["weight"]) for r in daily_rows if r["weight"] is not None]
    chart_calorie_labels = [r["entry_date"] for r in daily_rows if r["calories"] is not None]
    chart_calories = [int(r["calories"]) for r in daily_rows if r["calories"] is not None]

    return render_template(
        "index.html",
        profile=profile,
        projection=projection,
        weekly_rows=weekly_rows,
        daily_rows=daily_rows,
        current_weight=current_weight,
        current_bmi=current_bmi,
        current_bmr=current_bmr,
        avg_daily_calories=avg_daily_calories,
        selected_tdee=selected_tdee,
        estimated_daily_deficit=estimated_daily_deficit,
        estimated_weekly_loss=estimated_weekly_loss,
        chart_labels=chart_labels,
        chart_weights=chart_weights,
        chart_calorie_labels=chart_calorie_labels,
        chart_calories=chart_calories,
    )


@app.route("/profile", methods=["GET", "POST"])
def profile_view() -> str:
    db = get_db()
    profile = get_profile()
    if request.method == "POST":
        age = int(request.form["age"])
        start_date = request.form["start_date"]
        start_weight = float(request.form["start_weight"])
        height_inches = float(request.form["height_inches"])
        calorie_goal = int(request.form["calorie_goal"])
        activity_level = request.form.get("activity_level", DEFAULT_ACTIVITY_LEVEL)
        if activity_level not in ACTIVITY_MULTIPLIERS:
            activity_level = DEFAULT_ACTIVITY_LEVEL
        db.execute(
            """
            INSERT OR REPLACE INTO profile (id, age, start_date, start_weight, height_inches, calorie_goal, activity_level)
            VALUES (1, ?, ?, ?, ?, ?, ?)
            """,
            (age, start_date, start_weight, height_inches, calorie_goal, activity_level),
        )
        db.commit()
        flash("Profile updated.")
        return redirect(url_for("profile_view"))
    return render_template("profile.html", profile=profile)


@app.route("/weekly", methods=["GET", "POST"])
def weekly_view() -> str:
    db = get_db()
    if request.method == "POST":
        week_start = request.form["week_start"]
        actual_weight = request.form.get("actual_weight") or None
        weekly_calories = request.form.get("weekly_calories") or None
        note = request.form.get("note") or None
        db.execute(
            """
            INSERT INTO weekly_entries (week_start, actual_weight, weekly_calories, note, updated_at)
            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(week_start) DO UPDATE SET
                actual_weight = excluded.actual_weight,
                weekly_calories = excluded.weekly_calories,
                note = excluded.note,
                updated_at = CURRENT_TIMESTAMP
            """,
            (
                week_start,
                float(actual_weight) if actual_weight not in (None, "") else None,
                int(weekly_calories) if weekly_calories not in (None, "") else None,
                note,
            ),
        )
        db.commit()
        flash("Weekly entry saved.")
        return redirect(url_for("weekly_view"))

    rows = db.execute("SELECT * FROM weekly_entries ORDER BY week_start DESC").fetchall()
    return render_template("weekly.html", rows=rows, suggested_week_start=week_start_for(date.today()).isoformat())


@app.route("/daily", methods=["GET", "POST"])
def daily_view() -> str:
    db = get_db()
    if request.method == "POST":
        entry_date = request.form["entry_date"]
        weight = request.form.get("weight") or None
        calories = request.form.get("calories") or None
        db.execute(
            """
            INSERT INTO daily_entries (entry_date, weight, calories, updated_at)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(entry_date) DO UPDATE SET
                weight = excluded.weight,
                calories = excluded.calories,
                updated_at = CURRENT_TIMESTAMP
            """,
            (
                entry_date,
                float(weight) if weight not in (None, "") else None,
                int(calories) if calories not in (None, "") else None,
            ),
        )
        db.commit()
        flash("Daily entry saved.")
        return redirect(url_for("daily_view"))

    rows = db.execute("SELECT * FROM daily_entries ORDER BY entry_date DESC").fetchall()
    return render_template("daily.html", rows=rows, today_iso=date.today().isoformat())


@app.route("/daily/delete/<int:entry_id>", methods=["POST"])
def delete_daily_entry(entry_id: int) -> str:
    db = get_db()
    db.execute("DELETE FROM daily_entries WHERE id = ?", (entry_id,))
    db.commit()
    flash("Daily entry deleted.")
    return redirect(url_for("daily_view"))


@app.route("/weekly/delete/<int:entry_id>", methods=["POST"])
def delete_weekly_entry(entry_id: int) -> str:
    db = get_db()
    db.execute("DELETE FROM weekly_entries WHERE id = ?", (entry_id,))
    db.commit()
    flash("Weekly entry deleted.")
    return redirect(url_for("weekly_view"))


@app.route("/clear-data", methods=["POST"])
def clear_data() -> str:
    scope = request.form.get("scope", "all")
    db = get_db()

    if scope == "daily":
        db.execute("DELETE FROM daily_entries")
        flash("All daily entries cleared.")
    elif scope == "weekly":
        db.execute("DELETE FROM weekly_entries")
        flash("All weekly entries cleared.")
    elif scope == "all_entries":
        db.execute("DELETE FROM daily_entries")
        db.execute("DELETE FROM weekly_entries")
        flash("All daily and weekly entries cleared.")
    elif scope == "all_data":
        db.execute("DELETE FROM daily_entries")
        db.execute("DELETE FROM weekly_entries")
        db.execute("DELETE FROM profile")
        db.commit()
        seed_profile_from_defaults()
        flash("All tracker data cleared. Profile was reset from defaults or workbook import.")
        return redirect(url_for("index"))
    else:
        flash("Unknown clear-data option.")
        return redirect(url_for("index"))

    db.commit()
    next_page = request.form.get("next") or url_for("index")
    return redirect(next_page)


@app.route("/import-excel", methods=["POST"])
def import_excel_route() -> str:
    db = get_db()
    for xlsx_path in DEFAULT_XLSX_CANDIDATES:
        if xlsx_path.exists():
            db.execute("DELETE FROM weekly_entries")
            db.execute("DELETE FROM daily_entries")
            import_from_excel(db, xlsx_path)
            flash(f"Imported workbook from {xlsx_path.name}.")
            return redirect(url_for("index"))
    flash("No workbook file was found inside the container. Put the Excel file in ./data and try again.")
    return redirect(url_for("index"))


if __name__ == "__main__":
    init_db()
    app.run(host="0.0.0.0", port=1055, debug=True)
